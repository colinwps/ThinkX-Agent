"""
把一个 Excel 文件的所有 sheet 合并成一个 sheet
用法: python merge_sheets.py <input.xlsx> <output.xlsx>
"""
import sys

from openpyxl import Workbook, load_workbook


def merge(input_path: str, output_path: str):
    src = load_workbook(input_path, data_only=True)
    dst = Workbook()
    dst_ws = dst.active
    dst_ws.title = "merged"

    header_written = False
    total_rows = 0

    for sheet_name in src.sheetnames:
        ws = src[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        if not header_written:
            # 第一个 sheet 写表头 + 数据
            dst_ws.append(["_source_sheet", *rows[0]])
            header_written = True

        # 跳过表头, 写数据行
        for row in rows[1:]:
            dst_ws.append([sheet_name, *row])
            total_rows += 1

    dst.save(output_path)
    print(f"已合并 {len(src.sheetnames)} 个 sheet, 共 {total_rows} 行 -> {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("用法: python merge_sheets.py <input.xlsx> <output.xlsx>")
        sys.exit(1)
    merge(sys.argv[1], sys.argv[2])
